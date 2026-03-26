from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from codemap.models import ScanResult


_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4


def _apply_border(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).border = _THIN_BORDER


def _write_fields_tree(ws, fields, row: int, col_count: int, depth: int = 0) -> int:
    """Write JavaField tree with indentation. Returns next row."""
    indent = "  " * depth + ("ㄴ " if depth > 0 else "")
    for f in fields:
        ws.cell(row=row, column=1, value=f"{indent}{f.name}")
        ws.cell(row=row, column=2, value=f.type)
        ws.cell(row=row, column=3, value=f.comment)
        _apply_border(ws, row, col_count)
        row += 1
        if f.children:
            row = _write_fields_tree(ws, f.children, row, col_count, depth + 1)
    return row


def export_table_spec_xlsx(scan: ScanResult, output: Path) -> None:
    wb = openpyxl.Workbook()

    # -- Index sheet ("목차") --
    ws_index = wb.active
    ws_index.title = "목차"
    index_headers = ["No", "테이블명", "설명"]
    for ci, h in enumerate(index_headers, 1):
        ws_index.cell(row=1, column=ci, value=h)
    _style_header_row(ws_index, 1, len(index_headers))

    for ti, table in enumerate(scan.database.tables, 1):
        ws_index.cell(row=ti + 1, column=1, value=ti)
        ws_index.cell(row=ti + 1, column=2, value=table.name)
        ws_index.cell(row=ti + 1, column=3, value=table.comment)
        _apply_border(ws_index, ti + 1, len(index_headers))

    _auto_width(ws_index)

    # -- Per-table sheets --
    for table in scan.database.tables:
        # Build FK map per table to avoid cross-table contamination
        fk_map: dict[str, str] = {fk.column: fk.references for fk in table.foreignKeys}

        ws = wb.create_sheet(title=table.name)

        col_headers = ["No", "컬럼명", "타입", "PK", "FK", "Nullable", "설명"]
        for ci, h in enumerate(col_headers, 1):
            ws.cell(row=1, column=ci, value=h)
        _style_header_row(ws, 1, len(col_headers))

        row = 2
        for ci, col in enumerate(table.columns, 1):
            fk_ref = fk_map.get(col.name, "")
            ws.cell(row=row, column=1, value=ci)
            ws.cell(row=row, column=2, value=col.name)
            ws.cell(row=row, column=3, value=col.type)
            ws.cell(row=row, column=4, value="Y" if col.pk else "")
            ws.cell(row=row, column=5, value=fk_ref)
            ws.cell(row=row, column=6, value="Y" if col.nullable else "N")
            ws.cell(row=row, column=7, value=col.comment)
            _apply_border(ws, row, len(col_headers))
            row += 1

        # Index section
        if table.indexes:
            row += 1
            idx_headers = ["No", "인덱스명", "컬럼", "Unique"]
            for ci, h in enumerate(idx_headers, 1):
                ws.cell(row=row, column=ci, value=h)
            _style_header_row(ws, row, len(idx_headers))
            row += 1
            for ii, idx in enumerate(table.indexes, 1):
                ws.cell(row=row, column=1, value=ii)
                ws.cell(row=row, column=2, value=idx.name)
                ws.cell(row=row, column=3, value=", ".join(idx.columns))
                ws.cell(row=row, column=4, value="Y" if idx.unique else "N")
                _apply_border(ws, row, len(idx_headers))
                row += 1

        _auto_width(ws)

    wb.save(output)


def export_api_spec_xlsx(scan: ScanResult, output: Path) -> None:
    wb = openpyxl.Workbook()

    # --- 요약 시트 ---
    ws = wb.active
    ws.title = "엔드포인트 목록"
    headers = ["No", "메서드", "경로", "컨트롤러", "서비스", "반환 타입"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header_row(ws, 1, len(headers))

    for ei, ep in enumerate(scan.api.endpoints, 1):
        row = ei + 1
        ws.cell(row=row, column=1, value=ei)
        ws.cell(row=row, column=2, value=ep.method)
        ws.cell(row=row, column=3, value=ep.path)
        ws.cell(row=row, column=4, value=ep.controller)
        ws.cell(row=row, column=5, value=ep.service)
        ws.cell(row=row, column=6, value=ep.returnType)
        _apply_border(ws, row, len(headers))
    _auto_width(ws)

    # --- 엔드포인트별 상세 시트 ---
    for ei, ep in enumerate(scan.api.endpoints, 1):
        # Build short sheet name: "01_GET _api_users"
        raw_name = f"{ep.method} {ep.path}"
        for ch in r'/\?*[]:':
            raw_name = raw_name.replace(ch, "_")
        # Use index prefix + truncate to stay within 31 char limit
        prefix = f"{ei:02d}_"
        sheet_name = prefix + raw_name[:31 - len(prefix)]
        ws_ep = wb.create_sheet(title=sheet_name)
        row = 1

        # 기본 정보
        ws_ep.cell(row=row, column=1, value="메서드")
        ws_ep.cell(row=row, column=2, value=ep.method)
        row += 1
        ws_ep.cell(row=row, column=1, value="경로")
        ws_ep.cell(row=row, column=2, value=ep.path)
        row += 1
        ws_ep.cell(row=row, column=1, value="컨트롤러")
        ws_ep.cell(row=row, column=2, value=ep.controller)
        row += 1
        ws_ep.cell(row=row, column=1, value="서비스")
        ws_ep.cell(row=row, column=2, value=ep.service)
        row += 1
        ws_ep.cell(row=row, column=1, value="반환 타입")
        ws_ep.cell(row=row, column=2, value=ep.returnType)
        row += 2

        # 입력 파라미터
        query_params = [p for p in ep.params if p.annotation != "RequestBody"]
        if query_params:
            ws_ep.cell(row=row, column=1, value="입력 파라미터")
            ws_ep.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            param_headers = ["파라미터명", "타입", "어노테이션", "필수"]
            for ci, h in enumerate(param_headers, 1):
                ws_ep.cell(row=row, column=ci, value=h)
            _style_header_row(ws_ep, row, len(param_headers))
            row += 1
            for p in query_params:
                ws_ep.cell(row=row, column=1, value=p.name)
                ws_ep.cell(row=row, column=2, value=p.type)
                ws_ep.cell(row=row, column=3, value=f"@{p.annotation}")
                ws_ep.cell(row=row, column=4, value="Y" if p.required else "N")
                _apply_border(ws_ep, row, len(param_headers))
                row += 1
            row += 1

        # 요청 본문
        if ep.requestFields:
            ws_ep.cell(row=row, column=1, value="요청 본문")
            ws_ep.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            field_headers = ["필드명", "타입", "설명"]
            for ci, h in enumerate(field_headers, 1):
                ws_ep.cell(row=row, column=ci, value=h)
            _style_header_row(ws_ep, row, len(field_headers))
            row += 1
            for f in ep.requestFields:
                ws_ep.cell(row=row, column=1, value=f.name)
                ws_ep.cell(row=row, column=2, value=f.type)
                ws_ep.cell(row=row, column=3, value=f.comment)
                _apply_border(ws_ep, row, len(field_headers))
                row += 1
            row += 1

        # 응답
        if ep.responseFields:
            ws_ep.cell(row=row, column=1, value="응답")
            ws_ep.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            field_headers = ["필드명", "타입", "설명"]
            for ci, h in enumerate(field_headers, 1):
                ws_ep.cell(row=row, column=ci, value=h)
            _style_header_row(ws_ep, row, len(field_headers))
            row += 1
            row = _write_fields_tree(ws_ep, ep.responseFields, row, len(field_headers))

        _auto_width(ws_ep)

    wb.save(output)
