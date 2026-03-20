from pathlib import Path

import openpyxl

from codemap.models import (
    ScanResult, DatabaseSchema, Table, Column, ForeignKey, Index,
    ApiSchema, Endpoint, JavaField, Param,
)
from codemap.export.xlsx import export_table_spec_xlsx, export_api_spec_xlsx


def _sample_scan():
    return ScanResult(
        project="test",
        database=DatabaseSchema(tables=[
            Table(name="users", columns=[
                Column(name="id", type="BIGINT", pk=True, nullable=False),
                Column(name="email", type="VARCHAR(255)", nullable=False),
            ], foreignKeys=[ForeignKey(column="dept_id", references="departments.id")],
            indexes=[Index(name="idx_email", columns=["email"], unique=True)]),
            Table(name="departments", columns=[
                Column(name="id", type="BIGINT", pk=True, nullable=False),
                Column(name="name", type="VARCHAR(255)", nullable=False),
            ]),
        ]),
        api=ApiSchema(endpoints=[
            Endpoint(method="GET", path="/api/users", controller="UserController",
                     service="UserService", calls=["UserRepository.findAll"]),
        ]),
    )


def test_export_table_spec_xlsx(tmp_path):
    output = tmp_path / "table-spec.xlsx"
    export_table_spec_xlsx(_sample_scan(), output)
    assert output.exists()

    wb = openpyxl.load_workbook(output)
    assert "목차" in wb.sheetnames
    assert "users" in wb.sheetnames
    assert "departments" in wb.sheetnames


def test_export_table_spec_xlsx_index_sheet(tmp_path):
    output = tmp_path / "table-spec.xlsx"
    export_table_spec_xlsx(_sample_scan(), output)
    wb = openpyxl.load_workbook(output)
    ws = wb["목차"]
    assert ws.max_row >= 3


def test_export_table_spec_xlsx_table_sheet(tmp_path):
    output = tmp_path / "table-spec.xlsx"
    export_table_spec_xlsx(_sample_scan(), output)
    wb = openpyxl.load_workbook(output)
    ws = wb["users"]
    assert ws.max_row >= 3
    assert ws.cell(1, 2).value == "컬럼명"


def test_export_api_spec_xlsx(tmp_path):
    output = tmp_path / "api-spec.xlsx"
    export_api_spec_xlsx(_sample_scan(), output)
    assert output.exists()
    wb = openpyxl.load_workbook(output)
    assert "엔드포인트 목록" in wb.sheetnames


def test_export_api_spec_xlsx_detailed(tmp_path):
    """엔드포인트별 시트에 요청/응답 필드가 포함됨"""
    scan = ScanResult(
        project="test",
        api=ApiSchema(endpoints=[
            Endpoint(
                method="GET", path="/api/devices",
                controller="DeviceApiController", service="DeviceService",
                params=[Param(name="deviceType", type="String", annotation="RequestParam", required=False)],
                returnType="ApiResponse<List<Device>>",
                responseFields=[
                    JavaField(name="deviceId", type="Integer", comment="고유번호"),
                    JavaField(name="deviceName", type="String", comment="장비명"),
                ],
            ),
            Endpoint(
                method="POST", path="/api/users",
                controller="UserController", service="UserService",
                params=[Param(name="user", type="User", annotation="RequestBody")],
                returnType="ApiResponse<User>",
                requestFields=[
                    JavaField(name="email", type="String", comment="이메일"),
                ],
                responseFields=[
                    JavaField(name="id", type="Long", comment="고유번호"),
                ],
            ),
        ]),
    )
    output = tmp_path / "api-detail.xlsx"
    export_api_spec_xlsx(scan, output)
    wb = openpyxl.load_workbook(output)
    assert "엔드포인트 목록" in wb.sheetnames
    assert len(wb.sheetnames) >= 3
